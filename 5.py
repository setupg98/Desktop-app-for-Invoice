#!/usr/bin/env python3
"""
invoice_app_three_pages.py
- Company Details page (persist to company.json)
- Invoice page (customers, products, per-item discount%, tax%, totals)
- History page (SQLite invoices.db, open PDF, delete)
- PDF preview (opens system default viewer)
"""

import os
import json
import sqlite3
import webbrowser
import tempfile
from datetime import datetime
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from fpdf import FPDF
from PIL import Image, ImageTk
from openpyxl import Workbook

# ----------------- Paths & Setup -----------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_FILE = os.path.join(BASE_DIR, "invoices.db")
COMPANY_FILE = os.path.join(BASE_DIR, "company.json")
PDF_DIR = os.path.join(BASE_DIR, "invoices")
os.makedirs(PDF_DIR, exist_ok=True)

# ----------------- Database -----------------
conn = sqlite3.connect(DB_FILE)
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    address TEXT,
    phone TEXT,
    email TEXT
)''')

c.execute('''CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    price REAL NOT NULL,
    barcode TEXT
)''')

c.execute('''CREATE TABLE IF NOT EXISTS invoices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_number TEXT,
    customer_id INTEGER,
    subtotal REAL,
    taxes REAL,
    discount_percent REAL,
    final_total REAL,
    status TEXT,
    payment_method TEXT,
    date TEXT,
    notes TEXT,
    file_path TEXT,
    FOREIGN KEY(customer_id) REFERENCES customers(id)
)''')

c.execute('''CREATE TABLE IF NOT EXISTS invoice_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id INTEGER,
    product_id INTEGER,
    product_name TEXT,
    qty INTEGER,
    unit_price REAL,
    discount_percent REAL,
    tax_percent REAL,
    subtotal REAL,
    FOREIGN KEY(invoice_id) REFERENCES invoices(id)
)''')
conn.commit()

# ----------------- Company config -----------------
DEFAULT_COMPANY = {
    "name": "My Company",
    "address": "123 Business Road, City",
    "contact": "Phone: +92-XXXXXXXXX | Email: info@company.com",
    "footer": "Thank you for your business!",
    "signature": "Authorized Signature",
    "logo": "",
    "watermark": ""
}

def load_company():
    if os.path.exists(COMPANY_FILE):
        try:
            with open(COMPANY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return DEFAULT_COMPANY.copy()
    return DEFAULT_COMPANY.copy()

def save_company(cfg):
    with open(COMPANY_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

COMPANY = load_company()

# ----------------- Helpers -----------------
def fmt_currency(x):
    try:
        return f"PKR {float(x):.2f}"
    except Exception:
        return "PKR 0.00"

def generate_invoice_number():
    # INV-YYYYMMDD-#### (incremental daily)
    today = datetime.now().strftime("%Y%m%d")
    row = c.execute("SELECT COUNT(*) FROM invoices WHERE date LIKE ?", (today + "%",)).fetchone()
    n = (row[0] or 0) + 1
    return f"INV-{today}-{n:04d}"

# ----------------- PDF Generator -----------------
class InvoicePDF(FPDF):
    def header(self):
        # logo left, company info right
        if COMPANY.get("logo"):
            try:
                self.image(COMPANY["logo"], 10, 8, 30)
            except Exception:
                pass
        self.set_font("Arial", "B", 14)
        self.cell(0, 6, COMPANY.get("name", ""), ln=True, align="R")
        self.set_font("Arial", "", 10)
        self.cell(0, 5, COMPANY.get("address", ""), ln=True, align="R")
        self.cell(0, 5, COMPANY.get("contact", ""), ln=True, align="R")
        self.ln(6)
        # watermark (subtle)
        if COMPANY.get("watermark"):
            self.set_text_color(230,230,230)
            self.set_font("Arial","B",40)
            # attempt to place watermark in center-ish
            self.rotate(15)
            self.text(40, 120, COMPANY["watermark"])
            self.rotate(0)
            self.set_text_color(0,0,0)

    def footer(self):
        self.set_y(-30)
        self.set_font("Arial", "", 9)
        self.cell(0, 6, COMPANY.get("footer",""), ln=True, align="C")
        self.cell(0, 6, f"Page {self.page_no()}", align="R")

    # minimal rotate stub (fpdf2 supports transformations differently; keep safe)
    def rotate(self, deg):
        # This function is a harmless placeholder for watermark rotation,
        # not a true rotation implementation. It's left empty to avoid complex transforms.
        return

def create_invoice_pdf(invoice_number, customer, items, invoice_discount_pct, date_str, status, payment_method, notes, save_path):
    """
    items: list of dicts with keys:
      product_name, qty, unit_price, discount_percent, tax_percent
    """
    pdf = InvoicePDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"INVOICE: {invoice_number}", ln=True, align="C")
    pdf.ln(4)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 6, f"Date: {date_str}", ln=True)
    pdf.cell(0, 6, f"Payment Status: {status}    Method: {payment_method}", ln=True)
    pdf.ln(4)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 6, "Bill To:", ln=True)
    pdf.set_font("Arial", "", 10)
    pdf.multi_cell(0, 5, f"{customer.get('name','Guest')}\n{customer.get('address','')}\nPhone: {customer.get('phone','')}\nEmail: {customer.get('email','')}")
    pdf.ln(4)

    # table header
    headers = ["Product", "Qty", "Price", "Disc%", "Tax%", "Total"]
    w = [80, 18, 28, 22, 22, 30]
    pdf.set_fill_color(70,130,180)
    pdf.set_text_color(255,255,255)
    pdf.set_font("Arial","B",10)
    for i,h in enumerate(headers):
        pdf.cell(w[i], 8, h, border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0,0,0)
    pdf.set_font("Arial","",10)

    subtotal = 0.0
    tax_totals = {}

    for it in items:
        pname = str(it['product_name'])
        qty = int(it['qty'])
        unit = float(it['unit_price'])
        d_pct = float(it.get('discount_percent', 0.0))
        t_pct = float(it.get('tax_percent', 0.0))
        base = unit * qty
        discount_amount = base * (d_pct/100.0)
        taxable = base - discount_amount
        tax_amount = taxable * (t_pct/100.0)
        line_total = taxable + tax_amount

        # page break header repeat if necessary
        if pdf.get_y() > 250:
            pdf.add_page()
            pdf.set_fill_color(70,130,180)
            pdf.set_text_color(255,255,255)
            pdf.set_font("Arial","B",10)
            for i,h in enumerate(headers):
                pdf.cell(w[i], 8, h, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_text_color(0,0,0)
            pdf.set_font("Arial","",10)

        pdf.cell(w[0], 7, pname[:45], border=1)
        pdf.cell(w[1], 7, str(qty), border=1, align="C")
        pdf.cell(w[2], 7, f"PKR {unit:.2f}", border=1, align="R")
        pdf.cell(w[3], 7, f"{d_pct:.2f}", border=1, align="R")
        pdf.cell(w[4], 7, f"{t_pct:.2f}", border=1, align="R")
        pdf.cell(w[5], 7, f"PKR {line_total:.2f}", border=1, align="R")
        pdf.ln()

        subtotal += taxable
        if t_pct:
            tax_totals.setdefault(t_pct, 0.0)
            tax_totals[t_pct] += tax_amount

    pdf.ln(6)
    pdf.set_font("Arial","",10)
    for pct, amt in sorted(tax_totals.items()):
        pdf.cell(0,6, f"Tax ({pct:.2f}%): {fmt_currency(amt)}", ln=True, align="R")
    invoice_disc_amount = subtotal * (float(invoice_discount_pct or 0.0) / 100.0)
    pdf.cell(0,6, f"Invoice Discount ({invoice_discount_pct:.2f}%): {fmt_currency(invoice_disc_amount)}", ln=True, align="R")
    final_total = subtotal + sum(tax_totals.values()) - invoice_disc_amount
    pdf.set_font("Arial","B",12)
    pdf.cell(0,8, f"FINAL TOTAL: {fmt_currency(final_total)}", ln=True, align="R")
    pdf.ln(6)
    if notes:
        pdf.set_font("Arial","",10)
        pdf.multi_cell(0,6, "Notes: " + notes)
    pdf.ln(8)
    pdf.cell(0,6, COMPANY.get("signature","Authorized Signature"), ln=True)
    pdf.output(save_path)
    return final_total

# ----------------- GUI -----------------
root = Tk()
root.title("Invoice System — Company / Invoice / History")
root.geometry("1200x820")

nb = ttk.Notebook(root)
nb.pack(fill=BOTH, expand=True)

# ---------- Tab 1: Company Details ----------
tab_company = Frame(nb)
nb.add(tab_company, text="Company Details")

Label(tab_company, text="Company Name:").grid(row=0,column=0, sticky=W, padx=8, pady=6)
ent_company_name = Entry(tab_company, width=70)
ent_company_name.grid(row=0,column=1, padx=8, pady=6, sticky=W)
ent_company_name.insert(0, COMPANY.get("name",""))

Label(tab_company, text="Address:").grid(row=1,column=0, sticky=NW, padx=8)
txt_company_address = Text(tab_company, width=60, height=4)
txt_company_address.grid(row=1,column=1, padx=8, pady=6, sticky=W)
txt_company_address.insert("1.0", COMPANY.get("address",""))

Label(tab_company, text="Contact:").grid(row=2,column=0, sticky=W, padx=8)
ent_company_contact = Entry(tab_company, width=70)
ent_company_contact.grid(row=2,column=1, padx=8, pady=6, sticky=W)
ent_company_contact.insert(0, COMPANY.get("contact",""))

Label(tab_company, text="Footer text:").grid(row=3,column=0, sticky=W, padx=8)
ent_company_footer = Entry(tab_company, width=70)
ent_company_footer.grid(row=3,column=1, padx=8, pady=6, sticky=W)
ent_company_footer.insert(0, COMPANY.get("footer",""))

Label(tab_company, text="Signature text:").grid(row=4,column=0, sticky=W, padx=8)
ent_company_sig = Entry(tab_company, width=70)
ent_company_sig.grid(row=4,column=1, padx=8, pady=6, sticky=W)
ent_company_sig.insert(0, COMPANY.get("signature",""))

Label(tab_company, text="Watermark (optional):").grid(row=5,column=0, sticky=W, padx=8)
ent_company_wm = Entry(tab_company, width=40)
ent_company_wm.grid(row=5,column=1, padx=8, pady=6, sticky=W)
ent_company_wm.insert(0, COMPANY.get("watermark",""))

Label(tab_company, text="Logo:").grid(row=6,column=0, sticky=W, padx=8)
logo_frame = Frame(tab_company)
logo_frame.grid(row=6,column=1, sticky=W, padx=8, pady=6)
logo_path_var = StringVar(value=COMPANY.get("logo",""))
ent_logo_path = Entry(logo_frame, width=50, textvariable=logo_path_var)
ent_logo_path.pack(side=LEFT)
def choose_logo():
    p = filedialog.askopenfilename(filetypes=[("Image files","*.png;*.jpg;*.jpeg;*.bmp")])
    if p:
        logo_path_var.set(p)
        # show small preview
        try:
            img = Image.open(p); img.thumbnail((180,90))
            imgtk = ImageTk.PhotoImage(img)
            lbl_logo_preview.config(image=imgtk)
            lbl_logo_preview.image = imgtk
        except Exception:
            pass
Button(logo_frame, text="Browse", command=choose_logo).pack(side=LEFT, padx=6)

lbl_logo_preview = Label(tab_company)
lbl_logo_preview.grid(row=7,column=1, sticky=W, padx=8)
if COMPANY.get("logo") and os.path.exists(COMPANY.get("logo")):
    try:
        img = Image.open(COMPANY["logo"]); img.thumbnail((180,90))
        lbl_logo_preview.image = ImageTk.PhotoImage(img)
        lbl_logo_preview.config(image=lbl_logo_preview.image)
    except Exception:
        pass

def save_company():
    COMPANY["name"] = ent_company_name.get().strip()
    COMPANY["address"] = txt_company_address.get("1.0", END).strip()
    COMPANY["contact"] = ent_company_contact.get().strip()
    COMPANY["footer"] = ent_company_footer.get().strip()
    COMPANY["signature"] = ent_company_sig.get().strip()
    COMPANY["watermark"] = ent_company_wm.get().strip()
    COMPANY["logo"] = logo_path_var.get().strip()
    save_company(COMPANY)
    messagebox.showinfo("Saved", "Company details saved.")
Button(tab_company, text="Save Company Details", command=save_company).grid(row=8, column=1, sticky=W, padx=8, pady=12)

# ---------- Tab 2: Invoice ----------
tab_invoice = Frame(nb)
nb.add(tab_invoice, text="Create Invoice")

# Customer area
cust_frame = LabelFrame(tab_invoice, text="Customer", padx=8, pady=8)
cust_frame.pack(fill=X, padx=10, pady=6)

Label(cust_frame, text="Name:").grid(row=0,column=0, sticky=W)
entry_cust_name = Entry(cust_frame, width=30); entry_cust_name.grid(row=0,column=1,padx=6)
Label(cust_frame, text="Phone:").grid(row=0,column=2, sticky=W)
entry_cust_phone = Entry(cust_frame, width=20); entry_cust_phone.grid(row=0,column=3,padx=6)
Label(cust_frame, text="Email:").grid(row=1,column=0, sticky=W)
entry_cust_email = Entry(cust_frame, width=30); entry_cust_email.grid(row=1,column=1,padx=6)
Label(cust_frame, text="Address:").grid(row=2,column=0, sticky=NW)
entry_cust_address = Text(cust_frame, width=60, height=3); entry_cust_address.grid(row=2,column=1, columnspan=3, padx=6)

# quick load customers dropdown
Label(cust_frame, text="Select Customer:").grid(row=0,column=4, sticky=W, padx=(20,6))
cust_combo = ttk.Combobox(cust_frame, width=30)
cust_combo.grid(row=0,column=5)
def refresh_customers_combo():
    rows = c.execute("SELECT id,name FROM customers ORDER BY name").fetchall()
    cust_combo['values'] = [f"{r[0]} - {r[1]}" for r in rows]
refresh_customers_combo()
def load_customer(event=None):
    v = cust_combo.get().strip()
    if not v:
        return
    try:
        cid = int(v.split('-',1)[0].strip())
    except:
        return
    r = c.execute("SELECT name,address,phone,email FROM customers WHERE id=?", (cid,)).fetchone()
    if r:
        entry_cust_name.delete(0,END); entry_cust_name.insert(0,r[0])
        entry_cust_address.delete("1.0", END); entry_cust_address.insert("1.0", r[1] or "")
        entry_cust_phone.delete(0,END); entry_cust_phone.insert(0,r[2] or "")
        entry_cust_email.delete(0,END); entry_cust_email.insert(0,r[3] or "")
cust_combo.bind("<<ComboboxSelected>>", load_customer)

def save_customer():
    name = entry_cust_name.get().strip()
    if not name:
        messagebox.showerror("Error", "Customer name required")
        return
    addr = entry_cust_address.get("1.0",END).strip()
    phone = entry_cust_phone.get().strip()
    email = entry_cust_email.get().strip()
    c.execute("INSERT INTO customers (name,address,phone,email) VALUES (?,?,?,?)", (name,addr,phone,email))
    conn.commit()
    refresh_customers_combo()
    messagebox.showinfo("Saved", "Customer saved.")

Button(cust_frame, text="Save Customer", command=save_customer).grid(row=3, column=1, sticky=W, pady=6)

# Product area
prod_frame = LabelFrame(tab_invoice, text="Products", padx=8, pady=8)
prod_frame.pack(fill=X, padx=10, pady=6)

Label(prod_frame, text="Name:").grid(row=0,column=0, sticky=W)
entry_prod_name = Entry(prod_frame, width=30); entry_prod_name.grid(row=0,column=1,padx=6)
Label(prod_frame, text="Price:").grid(row=0,column=2, sticky=W)
entry_prod_price = Entry(prod_frame, width=12); entry_prod_price.grid(row=0,column=3,padx=6)
Label(prod_frame, text="Barcode (opt):").grid(row=0,column=4, sticky=W)
entry_prod_barcode = Entry(prod_frame, width=20); entry_prod_barcode.grid(row=0,column=5,padx=6)

def add_product_db():
    name = entry_prod_name.get().strip()
    try:
        price = float(entry_prod_price.get().strip())
    except:
        messagebox.showerror("Error","Invalid price")
        return
    barcode = entry_prod_barcode.get().strip()
    c.execute("INSERT INTO products (name,price,barcode) VALUES (?,?,?)", (name,price,barcode))
    conn.commit()
    entry_prod_name.delete(0,END); entry_prod_price.delete(0,END); entry_prod_barcode.delete(0,END)
    refresh_products_combo()
    messagebox.showinfo("Added", "Product added.")
Button(prod_frame, text="Add Product", command=add_product_db).grid(row=0,column=6,padx=6)

Label(prod_frame, text="Select product:").grid(row=1,column=0, sticky=W)
prod_combo = ttk.Combobox(prod_frame, width=70)
prod_combo.grid(row=1,column=1, columnspan=4, sticky=W, padx=6)
def refresh_products_combo():
    rows = c.execute("SELECT id,name,price,barcode FROM products ORDER BY name").fetchall()
    prod_combo['values'] = [f"{r[0]} - {r[1]} (PKR {r[2]:.2f})" for r in rows]
refresh_products_combo()

Label(prod_frame, text="Qty:").grid(row=1,column=5, sticky=W)
entry_qty = Entry(prod_frame, width=6); entry_qty.grid(row=1,column=6, sticky=W)
entry_qty.insert(0,"1")
Label(prod_frame, text="Item Disc %:").grid(row=2,column=0, sticky=W)
entry_item_disc = Entry(prod_frame, width=6); entry_item_disc.grid(row=2,column=1, sticky=W); entry_item_disc.insert(0,"0")
Label(prod_frame, text="Item Tax %:").grid(row=2,column=2, sticky=W)
entry_item_tax = Entry(prod_frame, width=6); entry_item_tax.grid(row=2,column=3, sticky=W); entry_item_tax.insert(0,"0")

def add_item_to_invoice():
    sel = prod_combo.get().strip()
    if not sel:
        messagebox.showerror("Error", "Select a product")
        return
    try:
        pid = int(sel.split('-',1)[0].strip())
    except:
        messagebox.showerror("Error", "Invalid product selection")
        return
    row = c.execute("SELECT name,price FROM products WHERE id=?", (pid,)).fetchone()
    if not row:
        messagebox.showerror("Not found", "Product not found in DB")
        return
    pname, price = row
    try:
        qty = int(entry_qty.get())
    except:
        messagebox.showerror("Error", "Quantity must be integer")
        return
    try:
        d = float(entry_item_disc.get())
        t = float(entry_item_tax.get())
    except:
        messagebox.showerror("Error", "Discount and Tax must be numbers")
        return
    base = price * qty
    disc_amt = base * (d/100.0)
    taxable = base - disc_amt
    tax_amt = taxable * (t/100.0)
    subtotal = taxable + tax_amt
    items_tree.insert('', END, values=(pid, pname, qty, f"{price:.2f}", f"{d:.2f}", f"{t:.2f}", f"{subtotal:.2f}"))
    compute_totals()

Button(prod_frame, text="Add to Invoice", command=add_item_to_invoice).grid(row=2,column=6,padx=6)

# items tree
items_frame = Frame(tab_invoice)
items_frame.pack(fill=BOTH, expand=True, padx=10, pady=6)
cols = ("prod_id","product","qty","unit_price","disc_pct","tax_pct","subtotal")
items_tree = ttk.Treeview(items_frame, columns=cols, show='headings', height=12)
for col,txt in [("prod_id","ID"),("product","Product"),("qty","Qty"),("unit_price","Unit Price"),
                ("disc_pct","Disc %"),("tax_pct","Tax %"),("subtotal","Subtotal")]:
    items_tree.heading(col, text=txt)
    items_tree.column(col, width=120 if col!="product" else 360)
items_tree.pack(side=LEFT, fill=BOTH, expand=True)
scroll_y = Scrollbar(items_frame, orient=VERTICAL, command=items_tree.yview)
scroll_y.pack(side=RIGHT, fill=Y)
items_tree.configure(yscrollcommand=scroll_y.set)

# bottom invoice controls
bottom_frame = Frame(tab_invoice)
bottom_frame.pack(fill=X, padx=10, pady=8)
Label(bottom_frame, text="Invoice-level Discount %:").pack(side=LEFT, padx=6)
entry_invoice_disc = Entry(bottom_frame, width=6); entry_invoice_disc.pack(side=LEFT); entry_invoice_disc.insert(0,"0")
Label(bottom_frame, text="    Payment Status:").pack(side=LEFT, padx=6)
status_var = StringVar(value="Unpaid")
OptionMenu(bottom_frame, status_var, "Unpaid","Paid","Partial").pack(side=LEFT)
Label(bottom_frame, text="    Payment Method:").pack(side=LEFT, padx=6)
pay_method_var = StringVar(value="Cash")
OptionMenu(bottom_frame, pay_method_var, "Cash","Card","Online").pack(side=LEFT)
Button(bottom_frame, text="Compute Totals", command=lambda: compute_totals()).pack(side=LEFT, padx=8)
Button(bottom_frame, text="Preview PDF", command=lambda: preview_pdf()).pack(side=LEFT, padx=8)
Button(bottom_frame, text="Save Invoice", command=lambda: save_invoice()).pack(side=LEFT, padx=8)
Button(bottom_frame, text="Remove Selected Item", command=lambda: remove_selected_item()).pack(side=LEFT, padx=8)
Label(bottom_frame, text="    ").pack(side=LEFT, padx=6)
lbl_subtotal = Label(bottom_frame, text="Subtotal: PKR 0.00", font=("Arial",10,"bold"))
lbl_subtotal.pack(side=RIGHT, padx=12)
lbl_final = Label(bottom_frame, text="Final Total: PKR 0.00", font=("Arial",12,"bold"))
lbl_final.pack(side=RIGHT, padx=12)

def remove_selected_item():
    sel = items_tree.selection()
    if not sel:
        return
    for s in sel:
        items_tree.delete(s)
    compute_totals()

def compute_totals():
    subtotal = 0.0
    taxes_total = 0.0
    for iid in items_tree.get_children():
        vals = items_tree.item(iid)['values']
        # vals: pid, pname, qty, unit_price, disc_pct, tax_pct, subtotal
        try:
            qty = int(vals[2]); unit = float(vals[3]); disc = float(vals[4]); t_pct = float(vals[5])
        except:
            continue
        base = unit * qty
        discount_amount = base * (disc/100.0)
        taxable = base - discount_amount
        tax_amount = taxable * (t_pct/100.0)
        subtotal += taxable
        taxes_total += tax_amount
    try:
        inv_disc_pct = float(entry_invoice_disc.get())
    except:
        inv_disc_pct = 0.0
    inv_disc_amount = subtotal * (inv_disc_pct/100.0)
    final_total = subtotal + taxes_total - inv_disc_amount
    lbl_subtotal.config(text=f"Subtotal: PKR {subtotal:.2f}")
    lbl_final.config(text=f"Final Total: PKR {final_total:.2f}")
    return subtotal, taxes_total, inv_disc_pct, final_total

def build_invoice_data_from_form():
    # customer dict
    cust = {
        "name": entry_cust_name.get().strip() or "Guest",
        "address": entry_cust_address.get("1.0",END).strip(),
        "phone": entry_cust_phone.get().strip(),
        "email": entry_cust_email.get().strip()
    }
    items = []
    for iid in items_tree.get_children():
        pid, pname, qty, unit_price, d_pct, t_pct, subtotal = items_tree.item(iid)['values']
        items.append({
            "product_name": pname,
            "qty": int(qty),
            "unit_price": float(unit_price),
            "discount_percent": float(d_pct),
            "tax_percent": float(t_pct)
        })
    return cust, items

def preview_pdf():
    cust, items = build_invoice_data_from_form()
    if not items:
        messagebox.showerror("Error", "Add items to invoice before preview")
        return
    inv_no = generate_invoice_number()
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inv_disc = float(entry_invoice_disc.get() or 0.0)
    tmp = os.path.join(PDF_DIR, f"{inv_no}_preview.pdf")
    create_invoice_pdf(inv_no, cust, items, inv_disc, date_str, status_var.get(), pay_method_var.get(), "", tmp)
    webbrowser.open_new(r"file:///" + os.path.abspath(tmp))

def save_invoice():
    cust, items = build_invoice_data_from_form()
    if not items:
        messagebox.showerror("Error","Invoice must contain at least one item")
        return
    # ensure customer exists (search by name+phone)
    c.execute("SELECT id FROM customers WHERE name=? AND phone=?", (cust['name'], cust['phone']))
    row = c.fetchone()
    if row:
        cust_id = row[0]
    else:
        c.execute("INSERT INTO customers (name,address,phone,email) VALUES (?,?,?,?)", (cust['name'], cust['address'], cust['phone'], cust['email']))
        conn.commit()
        cust_id = c.lastrowid
        refresh_customers_combo()
    subtotal, taxes_total, inv_disc_pct, final_total = compute_totals()
    inv_no = generate_invoice_number()
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    notes = ""  # could add notes field
    save_path = os.path.join(PDF_DIR, f"{inv_no}.pdf")
    # generate pdf
    final_total_calc = create_invoice_pdf(inv_no, cust, items, inv_disc_pct, date_str, status_var.get(), pay_method_var.get(), notes, save_path)
    # save invoice record
    c.execute("""INSERT INTO invoices (invoice_number, customer_id, subtotal, taxes, discount_percent, final_total, status, payment_method, date, notes, file_path)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (inv_no, cust_id, subtotal, taxes_total, inv_disc_pct, final_total_calc, status_var.get(), pay_method_var.get(), date_str, notes, save_path))
    conn.commit()
    invoice_id = c.lastrowid
    for it in items:
        # try find product id
        rowp = c.execute("SELECT id FROM products WHERE name=?", (it['product_name'],)).fetchone()
        pid = rowp[0] if rowp else None
        base = it['unit_price'] * it['qty']
        disc_amt = base * (it['discount_percent']/100.0)
        taxable = base - disc_amt
        tax_amt = taxable * (it['tax_percent']/100.0)
        line_subtotal = taxable + tax_amt
        c.execute("""INSERT INTO invoice_items (invoice_id, product_id, product_name, qty, unit_price, discount_percent, tax_percent, subtotal)
                     VALUES (?,?,?,?,?,?,?,?)""", (invoice_id, pid, it['product_name'], it['qty'], it['unit_price'], it['discount_percent'], it['tax_percent'], line_subtotal))
    conn.commit()
    messagebox.showinfo("Saved", f"Invoice saved and PDF generated:\n{save_path}")
    # clear items
    items_tree.delete(*items_tree.get_children())
    compute_totals()
    refresh_history()

# ---------- Tab 3: History ----------
tab_history = Frame(nb)
nb.add(tab_history, text="History")

hist_frame = Frame(tab_history); hist_frame.pack(fill=BOTH, expand=True, padx=8, pady=8)
hist_tree = ttk.Treeview(hist_frame, columns=("id","invoice","customer","date","status","total"), show='headings')
for col,w in [("id",60),("invoice",160),("customer",260),("date",160),("status",90),("total",120)]:
    hist_tree.heading(col, text=col.title())
    hist_tree.column(col, width=w)
hist_tree.pack(fill=BOTH, expand=True, side=LEFT)
hist_scroll = Scrollbar(hist_frame, command=hist_tree.yview); hist_scroll.pack(side=RIGHT, fill=Y)
hist_tree.configure(yscrollcommand=hist_scroll.set)

def refresh_history():
    for r in hist_tree.get_children(): hist_tree.delete(r)
    rows = c.execute("""SELECT i.id, i.invoice_number, IFNULL(c.name,'Guest'), i.date, i.status, i.final_total
                        FROM invoices i LEFT JOIN customers c ON i.customer_id=c.id ORDER BY i.id DESC""").fetchall()
    for r in rows:
        hist_tree.insert('', END, values=(r[0], r[1], r[2], r[3], r[4], f"PKR {float(r[5]):.2f}" if r[5] is not None else "PKR 0.00"))
refresh_history()

def open_invoice_file(event=None):
    sel = hist_tree.selection()
    if not sel:
        return
    iid = hist_tree.item(sel[0])['values'][0]
    row = c.execute("SELECT file_path FROM invoices WHERE id=?", (iid,)).fetchone()
    if row and row[0] and os.path.exists(row[0]):
        webbrowser.open_new(r"file:///" + os.path.abspath(row[0]))
    else:
        messagebox.showerror("Not found", "PDF file not found for this invoice (maybe deleted).")

hist_tree.bind("<Double-1>", open_invoice_file)

def delete_selected_invoice():
    sel = hist_tree.selection()
    if not sel:
        messagebox.showerror("Select", "Select an invoice to delete")
        return
    iid = hist_tree.item(sel[0])['values'][0]
    if not messagebox.askyesno("Confirm", "Delete selected invoice and its PDF?"):
        return
    row = c.execute("SELECT file_path FROM invoices WHERE id=?", (iid,)).fetchone()
    if row and row[0] and os.path.exists(row[0]):
        try:
            os.remove(row[0])
        except Exception:
            pass
    c.execute("DELETE FROM invoice_items WHERE invoice_id=?", (iid,))
    c.execute("DELETE FROM invoices WHERE id=?", (iid,))
    conn.commit()
    refresh_history()
    messagebox.showinfo("Deleted", "Invoice removed.")

Button(tab_history, text="Refresh", command=refresh_history).pack(side=LEFT, padx=8, pady=6)
Button(tab_history, text="Open PDF (double-click)", command=open_invoice_file).pack(side=LEFT, padx=8)
Button(tab_history, text="Delete Selected", command=delete_selected_invoice).pack(side=LEFT, padx=8)

# ---------- Excel export (history selection) ----------
def export_selected_to_excel():
    sel = hist_tree.selection()
    if not sel:
        messagebox.showerror("Select", "Select at least one invoice to export")
        return
    ids = [hist_tree.item(s)['values'][0] for s in sel]
    path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
    if not path:
        return
    wb = Workbook(); ws = wb.active
    ws.title = "Invoices"
    ws.append(["Invoice ID","Invoice No","Customer","Date","Status","Final Total","File Path"])
    for iid in ids:
        row = c.execute("SELECT i.id, i.invoice_number, IFNULL(c.name,'Guest'), i.date, i.status, i.final_total, i.file_path FROM invoices i LEFT JOIN customers c ON i.customer_id=c.id WHERE i.id=?", (iid,)).fetchone()
        ws.append(list(row))
    wb.save(path)
    messagebox.showinfo("Saved", f"Exported to {path}")

Button(tab_history, text="Export Selected to Excel", command=export_selected_to_excel).pack(side=LEFT, padx=8)

# ---------- initial refresh ----------
refresh_customers_combo()
refresh_products_combo()
refresh_history()

# ----------------- Run -----------------
root.mainloop()
